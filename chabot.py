import tkinter as tk
from tkinter import messagebox
from openpyxl import Workbook, load_workbook
import os

# Function to save responses
def save_responses():
    user_name = name_entry.get()
    user_age = age_entry.get()
    user_email = email_entry.get()
    
    # Your Excel saving logic here (you can use either the appending or new-sheet method discussed before)
    filename = "chatbot_responses.xlsx"
    if os.path.exists(filename):
        wb = load_workbook(filename)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Name", "Age", "Email"])
    ws.append([user_name, user_age, user_email])
    wb.save(filename)
    
    # Show a message box to confirm data saved
    messagebox.showinfo("Success", "Your responses have been saved.")
    
    # Clear the entry fields for next input
    name_entry.delete(0, tk.END)
    age_entry.delete(0, tk.END)
    email_entry.delete(0, tk.END)

# Set up the GUI
root = tk.Tk()
root.title("Chatbot")

# Create and pack the widgets
tk.Label(root, text="What is your name?").pack()
name_entry = tk.Entry(root)
name_entry.pack()

tk.Label(root, text="What is your age?").pack()
age_entry = tk.Entry(root)
age_entry.pack()

tk.Label(root, text="What is your email address?").pack()
email_entry = tk.Entry(root)
email_entry.pack()

submit_button = tk.Button(root, text="Submit", command=save_responses)
submit_button.pack()

# Start the GUI event loop
root.mainloop()